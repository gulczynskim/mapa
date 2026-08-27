"""
Convert the Ministry of Health (BASiW) sickness-absence files into the
compact frontend format.

Source: 8 password-protected Excel files in data/Absencje/, released to the
user on request (MZ Departament Analiz i Strategii, ref. AST.461.50.2026.BA,
20.08.2026), one per year 2017-2024. Original ZUS data, redistributed by MZ
as part of Mapy potrzeb zdrowotnych. Files are AES-encrypted (ECMA-376
agile); the password sits alongside them in data/Absencje/Hasło, and
data/Absencje/ is gitignored -- 161 MB of source that never belongs in the
repo.

Each file is one flat table, ~480k rows: Rok, Województwo, Powiat, Płeć,
Klasyfikacje, Kategoria, ICD10_KOD, ICD10_NAZWA, Liczba dni absencji,
Liczba zaświadczeń. The diagnosis dimension is strictly nested three deep
(rozdziały -> kategorie -> kody ICD-10, verified on 2017: 21/261/2007, every
category in exactly one chapter, every code in exactly one category).

The chapter set is NOT constant across years: 2017-2019 have 21, and from
2020 a 22nd appears ("Kody do celów specjalnych", ICD-10 chapter XXII --
the COVID-19 codes). Chapters are therefore matched by name against
CHAPTER_KEYS and an unrecognised one aborts the run rather than being
silently folded into a total nobody declared.

Only LEVEL 1 (Klasyfikacje) is aggregated out to the site -- the frontend
gives each variable one categorical slot besides Miara, so a three-level
taxonomy has to be projected onto one level, and chapters are the level
that survives it (short labels, the level BASiW itself publishes at). The
finer levels stay in these source files for offline analysis.

Two synthetic totals are emitted alongside the 21 chapters:
  "ogolem"          -- every chapter summed
  "ogolem_bez_ciazy" -- every chapter except "Ciąża, poród i okres połogu"
The pregnancy chapter is a third of all women's absence days and is
structurally male-empty, so it flips the headline gender ratio: in 2017 the
K/M ratio of absence days is 1.37 with it and 0.91 without. One combined
total on its own would mislead whichever way it was computed.

Powiat identification: the source carries names only, no TERYT. Matched on
(województwo, normalised powiat name) -> data/powiaty.json's own JPT_KOD_JE,
using the voivodeship to disambiguate names that repeat nationally
(grodziski, brzeski, ...). Verified 380/380 for 2017, and asserted per year
here -- the script refuses to write a partial file rather than silently
dropping counties.

Measures:
  dni              -- Liczba dni absencji
  zaswiadczenia    -- Liczba zaświadczeń
  dlugosc_srednia  -- dni / zaświadczenia, average length of one certificate
The "_srednia" suffix is load-bearing: isRateMeasure() (app.js,
audit_data.py, export_unit_csv.py) keys off it to drop %kobiet/%mężczyzn
for that measure, which don't decompose for an average.
"""

import hashlib
import io
import json
import os
import struct
import sys
import xml.etree.ElementTree as ET
from base64 import b64decode

import olefile
import openpyxl
from cryptography.hazmat.primitives.ciphers import Cipher, algorithms, modes

SRC_DIR = "../data/Absencje"
OUT_DIR = "../data"
OUT_PATH = os.path.join(OUT_DIR, "absencje.json")
YEARS = range(2017, 2025)

PREGNANCY_CHAPTER = "Ciąża, poród i okres połogu"

# Stable slug per ICD-10 chapter. Keyed on the source's own Klasyfikacje
# text so a wording change in a future edition fails loudly here (see
# chapter_key) instead of silently minting a new ageGroup nobody declared
# in variables.js.
CHAPTER_KEYS = {
    "Wybrane choroby zakaźne i pasożytnicze": "zakazne",
    "Nowotwory": "nowotwory",
    "Choroby krwi i narządów krwiotwórczych oraz wybrane choroby przebiegające z udziałem mechanizmów immunologicznych": "krwi",
    "Zaburzenia wydzielania wewnętrznego, stanu odżywienia i przemian metabolicznych": "wydzielania",
    "Zaburzenia psychiczne i zaburzenia zachowania": "psychiczne",
    "Choroby układu nerwowego": "nerwowy",
    "Choroby oka i przydatków oka": "oko",
    "Choroby ucha i wyrostka sutkowatego": "ucho",
    "Choroby układu krążenia": "krazenie",
    "Choroby układu oddechowego": "oddechowy",
    "Choroby układu pokarmowego": "pokarmowy",
    "Choroby skóry i tkanki podskórnej": "skora",
    "Choroby układu mięśniowo-szkieletowego i tkanki łącznej": "miesniowo_szkieletowy",
    "Choroby układu moczowo-płciowego": "moczowo_plciowy",
    "Ciąża, poród i okres połogu": "ciaza",
    "Wybrane stany rozpoczynające się w okresie okołoporodowym": "okoloporodowe",
    "Wady rozwojowe wrodzone, zniekształcenia i aberracje chromosomowe": "wady_wrodzone",
    "Objawy, cechy chorobowe oraz nieprawidłowe wyniki badań klinicznych i laboratoryjnych niesklasyfikowane gdzie indziej": "objawy",
    "Urazy, zatrucia i inne określone skutki działania czynników zewnętrznych": "urazy",
    "Zewnętrzne przyczyny zachorowania i zgonu": "przyczyny_zewnetrzne",
    "Czynniki wpływające na stan zdrowia i kontakt ze służbą zdrowia": "czynniki_zdrowotne",
    # ICD-10 chapter XXII, absent from the 2017-2019 files and present from
    # 2020 on -- in this data it is essentially COVID-19 (U07/U08/U09). Years
    # before 2020 legitimately carry no value for it, which the frontend
    # renders as "brak danych".
    "Kody do celów specjalnych": "cele_specjalne",
}

SEX_KEYS = {"K": "k", "M": "m"}


# --- ECMA-376 agile decryption -------------------------------------------
# msoffcrypto-tool isn't installed and these files can't be read without it;
# cryptography + olefile already are, so the ~40 lines it takes are inlined
# rather than adding a dependency for one script.

_BLK_IN = bytes([0xFE, 0xA7, 0xD2, 0x76, 0x3B, 0x4B, 0x9E, 0x79])
_BLK_VAL = bytes([0xD7, 0xAA, 0x0F, 0x6D, 0x30, 0x61, 0x34, 0x4E])
_BLK_KEY = bytes([0x14, 0x6E, 0x0B, 0xE7, 0xAB, 0xAC, 0xD0, 0xD6])
_NS = {"p": "http://schemas.microsoft.com/office/2006/keyEncryptor/password"}


def _hash(alg, *parts):
    h = hashlib.new(alg)
    for p in parts:
        h.update(p)
    return h.digest()


def _derive(password, salt, spin, alg, block, key_bytes):
    """Password -> key: salted hash, spun spinCount times, then bound to one
    of the three fixed block keys (verifier input / verifier value / key)."""
    h = _hash(alg, salt, password.encode("utf-16-le"))
    for i in range(spin):
        h = _hash(alg, struct.pack("<I", i), h)
    return (_hash(alg, h, block) + b"\x36" * key_bytes)[:key_bytes]


def _aes_cbc(key, iv, data):
    dec = Cipher(algorithms.AES(key), modes.CBC(iv)).decryptor()
    return dec.update(data) + dec.finalize()


def decrypt_xlsx(path, password):
    """Returns the decrypted .xlsx bytes of an ECMA-376 agile-encrypted
    workbook. Raises SystemExit on a wrong password rather than handing back
    garbage that openpyxl would fail on far from the cause."""
    ole = olefile.OleFileIO(path)
    xml = ET.fromstring(ole.openstream("EncryptionInfo").read()[8:])
    enc = xml.find(".//p:encryptedKey", _NS).attrib
    alg = {"SHA512": "sha512", "SHA256": "sha256", "SHA1": "sha1"}[enc["hashAlgorithm"].upper()]
    spin = int(enc["spinCount"])
    key_bytes = int(enc["keyBits"]) // 8
    salt = b64decode(enc["saltValue"])

    verifier = _aes_cbc(_derive(password, salt, spin, alg, _BLK_IN, key_bytes), salt, b64decode(enc["encryptedVerifierHashInput"]))
    expected = _aes_cbc(_derive(password, salt, spin, alg, _BLK_VAL, key_bytes), salt, b64decode(enc["encryptedVerifierHashValue"]))
    digest = _hash(alg, verifier)
    if digest != expected[: len(digest)]:
        raise SystemExit(f"{os.path.basename(path)}: wrong password")

    secret = _aes_cbc(_derive(password, salt, spin, alg, _BLK_KEY, key_bytes), salt, b64decode(enc["encryptedKeyValue"]))[:key_bytes]
    key_data = xml.find(".//{*}keyData").attrib
    data_salt = b64decode(key_data["saltValue"])
    block_size = int(key_data["blockSize"])

    packed = ole.openstream("EncryptedPackage").read()
    total = struct.unpack("<Q", packed[:8])[0]
    out, pos, block = io.BytesIO(), 8, 0
    while pos < len(packed):
        iv = _hash(alg, data_salt, struct.pack("<I", block))[:block_size]
        out.write(_aes_cbc(secret, iv, packed[pos:pos + 4096]))
        pos += 4096
        block += 1
    return out.getvalue()[:total]


# --- powiat name -> TERYT -------------------------------------------------

def normalise_powiat(name):
    """The source writes powiats bare and adjectival ("kamiennogórski") or
    city-prefixed ("m. Wrocław"); powiaty.json writes them as "powiat
    kamiennogórski" / "Wrocław". Strip both conventions down to a common
    form. ł -> l because the two sources disagree on it in a couple of city
    names."""
    s = name.lower().strip()
    for prefix in ("powiat ", "m. st. ", "m. ", "miasto "):
        if s.startswith(prefix):
            s = s[len(prefix):]
    return s.replace("ł", "l")


def build_teryt_index():
    """(województwo 2-digit code, normalised powiat name) -> 4-digit TERYT.
    Voivodeship-scoped because ~20 powiat names repeat across the country."""
    powiaty = json.load(open(os.path.join(OUT_DIR, "powiaty.json"), encoding="utf-8"))
    wojewodztwa = json.load(open(os.path.join(OUT_DIR, "wojewodztwa.json"), encoding="utf-8"))
    woj_codes = {f["properties"]["JPT_NAZWA_"].lower(): f["properties"]["JPT_KOD_JE"] for f in wojewodztwa["features"]}
    index = {}
    for f in powiaty["features"]:
        teryt = f["properties"]["JPT_KOD_JE"]
        index[(teryt[:2], normalise_powiat(f["properties"]["JPT_NAZWA_"]))] = teryt
    return woj_codes, index


def chapter_key(text):
    try:
        return CHAPTER_KEYS[text]
    except KeyError:
        raise SystemExit(
            f"Unknown ICD-10 chapter {text!r} -- the source's Klasyfikacje wording changed. "
            "Add it to CHAPTER_KEYS and declare the matching ageGroup in variables.js/variables.en.js."
        )


# --- conversion -----------------------------------------------------------

def read_year(path, password, woj_codes, teryt_index):
    """One year's file -> {(teryt, chapter_key, sex): [days, certificates]},
    summed up from the ICD-code rows. Rows are streamed, not materialised --
    each file is ~480k rows."""
    wb = openpyxl.load_workbook(io.BytesIO(decrypt_xlsx(path, password)), read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)

    header = next(rows)
    expected = ("Rok", "Województwo", "Powiat", "Płeć", "Klasyfikacje", "Kategoria", "ICD10_KOD", "ICD10_NAZWA", "Liczba dni absencji", "Liczba zaświadczeń")
    if tuple(header) != expected:
        raise SystemExit(f"{os.path.basename(path)}: unexpected columns {header!r}")

    totals = {}
    seen_units = set()
    seen_chapters = set()
    special_codes = {}
    for row in rows:
        _, woj, powiat, sex, chapter, _, icd_code, icd_name, days, certs = row
        seen_chapters.add(chapter)
        if chapter == "Kody do celów specjalnych":
            special_codes[icd_code] = special_codes.get(icd_code, [icd_name, 0])
            special_codes[icd_code][1] += days or 0
        woj_code = woj_codes.get((woj or "").lower())
        teryt = teryt_index.get((woj_code, normalise_powiat(powiat or "")))
        if teryt is None:
            raise SystemExit(f"{os.path.basename(path)}: no TERYT for {woj!r}/{powiat!r}")
        seen_units.add(teryt)

        sex_key = SEX_KEYS.get(sex)
        if sex_key is None:
            raise SystemExit(f"{os.path.basename(path)}: unexpected Płeć value {sex!r}")

        slot = totals.setdefault((teryt, chapter_key(chapter), sex_key), [0, 0])
        slot[0] += days or 0
        slot[1] += certs or 0

    if len(seen_units) != 380:
        raise SystemExit(f"{os.path.basename(path)}: {len(seen_units)} powiaty, expected 380")
    return totals, seen_chapters, special_codes


def emit(totals, year, chapter_keys, all_teryts, out):
    """Fold one year's chapter totals into the output tree, adding the two
    synthetic totals and deriving the average-length measure.

    The full teryt x chapter grid is materialised, because the source is
    ragged: it carries a row only where certificates were actually issued.
    An absent (powiat, chapter, sex) combination means nobody in that group
    took leave for that chapter -- a real 0, not missing data. Left as null
    it would knock out Różnica/Proporcja/%kobiet for the whole slice, which
    matters most exactly where the answer is most interesting: men have no
    rows at all under "Ciąża, poród i okres połogu", so every county's
    pregnancy map would have rendered "brak danych" instead of the female
    total it should show.

    Only chapters actually present in THIS year's file are materialised.
    Chapter XXII is genuinely absent before 2020 rather than zero, so it
    stays null there and the frontend shows "brak danych"."""
    preg = CHAPTER_KEYS[PREGNANCY_CHAPTER]
    rolled = dict(totals)
    for (teryt, chapter, sex), (days, certs) in totals.items():
        for synthetic in ("ogolem",) + (() if chapter == preg else ("ogolem_bez_ciazy",)):
            slot = rolled.setdefault((teryt, synthetic, sex), [0, 0])
            slot[0] += days
            slot[1] += certs

    groups = sorted(chapter_keys) + ["ogolem", "ogolem_bez_ciazy"]
    year_str = str(year)
    for teryt in all_teryts:
        year_slot = out.setdefault(teryt, {}).setdefault(year_str, {})
        for group in groups:
            days = {s: rolled.get((teryt, group, s), (0, 0))[0] for s in ("k", "m")}
            certs = {s: rolled.get((teryt, group, s), (0, 0))[1] for s in ("k", "m")}
            days["t"] = days["k"] + days["m"]
            certs["t"] = certs["k"] + certs["m"]
            year_slot[f"{group}__dni"] = dict(days)
            year_slot[f"{group}__zaswiadczenia"] = dict(certs)
            # No certificates means no meaningful average -- null, not 0, so
            # it reads as "brak danych" rather than "sick leaves here last
            # no time at all". The total is total days over total
            # certificates, NOT the mean of the two sexes' averages, which
            # would weight a county's men and women equally regardless of
            # how many of each were actually off sick.
            year_slot[f"{group}__dlugosc_srednia"] = {
                s: (round(days[s] / certs[s], 1) if certs[s] else None) for s in ("t", "k", "m")
            }


def main():
    password_path = os.path.join(SRC_DIR, "Hasło")
    if not os.path.exists(password_path):
        raise SystemExit(f"{password_path} not found -- it holds the password MZ sent with the files")
    password = open(password_path, encoding="utf-8").read().strip()

    woj_codes, teryt_index = build_teryt_index()
    all_teryts = sorted(set(teryt_index.values()))
    out = {}
    for year in YEARS:
        path = os.path.join(SRC_DIR, f"Plik zusdemo_{year}_absencja_plec_haslo.xlsx")
        if not os.path.exists(path):
            raise SystemExit(f"missing {path}")
        totals, chapters, special = read_year(path, password, woj_codes, teryt_index)
        emit(totals, year, {chapter_key(c) for c in chapters}, all_teryts, out)
        # The chapter count varies by year (COVID from 2020), so report it
        # rather than assuming -- a silent change here would quietly reshape
        # the "Ogółem" totals.
        note = ""
        if special:
            top = sorted(special.values(), key=lambda v: -v[1])[:2]
            note = "  [Kody do celów specjalnych: " + ", ".join(f"{name} {days} dni" for name, days in top) + "]"
        print(f"  {year}: ok, {len(chapters)} rozdziałów{note}", flush=True)

    with open(OUT_PATH, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, separators=(",", ":"))
    size_mb = os.path.getsize(OUT_PATH) / 1e6
    print(f"absencje: {len(out)} powiatów x {len(YEARS)} lat -> {OUT_PATH} ({size_mb:.1f} MB)")


if __name__ == "__main__":
    sys.exit(main())
