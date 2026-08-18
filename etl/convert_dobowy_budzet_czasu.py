"""
Dobowy budżet czasu ludności w 2023 r. (GUS Time Use Survey), województwo
level, wg płci -- source: user-supplied Excel file "Załącznik_Budżet Czasu
Ludności 2023 wg województw i płci.xlsx" (GUS Tablica 9, from
https://stat.gov.pl/obszary-tematyczne/warunki-zycia/dochody-wydatki-i-warunki-zycia-ludnosci/dobowy-budzet-czasu-ludnosci-w-2023r-,35,1.html),
NOT a BDL fetch -- a one-off 2023 survey, not annual/monthly data.

Sheet layout: three sheets ("Ogółem"/"Kobiety"/"Mężczyźni" -> t/m/k
directly, no computation needed) each containing THREE stacked tables:
  1. "PRZECIĘTNY CZAS TRWANIA CZYNNOŚCI" (mean time spent in the activity,
     averaged across the WHOLE population, zero-filled for non-participants)
  2. "PRZECIĘTNY CZAS WYKONYWANIA CZYNNOŚCI" (mean participation time,
     averaged only among people who did that specific activity that day)
  3. "PRZECIĘTNY ODSETEK OSÓB WYKONUJĄCYCH CZYNNOŚCI" (participation rate, %)
Only tables 1 and 2 are used here, per explicit user request -- table 3
(participation rate) is a different kind of measure entirely and wasn't
asked for.

Values are stored in the file as "H.MM" notation (e.g. 11.22 means 11h22m,
NOT 11.22 decimal hours -- confirmed by summing a category's own components
in minutes and matching the parent's own "H.MM" figure). Converted here to
plain integer MINUTES, both to avoid the ambiguity of displaying something
that looks like "H.MM" again but isn't, and because whole minutes read more
naturally than fractional decimal hours for most of these values.

Three-level category hierarchy (L1 = ALL-CAPS heading, L2 = its direct
sub-items, L3 = further "w tym ..." detail) -- only L1 and L2 are kept, per
explicit user request. Extracting the L1/L2 split correctly needed several
rounds of correction against the live file (see chat) -- summarized here:
  - L1 detection: the heading is all-caps EXCEPT single-letter Polish
    conjunctions/prepositions (e.g. "PRACE DOMOWE i OPIEKA..." keeps a
    lowercase "i" by Polish typographic convention), which a naive
    `.upper() == self` check misses entirely.
  - A bare "w tym:" header row (no value of its own) does NOT by itself
    mean "everything below is one level deeper" -- what it introduces
    depends on what row came immediately before it: following an L1 row,
    the header introduces genuine L2 items; following an L2 row, it
    introduces L3 detail of THAT item specifically, and the L3 block ends
    (reverting to L2) as soon as the running sum reaches/exceeds the L2
    parent's own value -- confirmed necessary because a later, unrelated
    L2 sibling (e.g. "Oglądanie telewizji i filmów", 124 min) can otherwise
    look like it fits under a small-value "w tym:" block by sheer
    arithmetic headroom if the cutoff is too loose.
  - Two manual overrides, per explicit user instruction (not derivable from
    the file's own structure/arithmetic alone):
    * PRACA ZAWODOWA's only sub-line ("w tym praca główna i dodatkowa") is
      NOT promoted to L2 -- left with zero L2 children.
    * DOJAZDY's three "Dojazdy (dojścia) do X" items are promoted to be
      DIRECT L2 children of DOJAZDY, skipping over the intermediate
      "w tym dojazdy i dojścia" line entirely (which would otherwise place
      them at L3, one level too deep to keep).

Voivodeship columns: only the 16 real voivodeships are used (matching
data/wojewodztwa.json's own 16 features exactly) -- the file's "Polska
ogółem" column (national total) and its "w tym: region warszawski
stołeczny" / "region mazowiecki" sub-split of mazowieckie (an EU NUTS-2
breakdown with no matching shape in this project's boundary layers) are
both deliberately skipped, per explicit user decision -- mazowieckie is
stored as the single whole-voivodeship figure the file already publishes
in its own "mazowieckie" column, not derived from the sub-split.
"""

import json
import os

import openpyxl

SRC = "../data/Załącznik_Budżet Czasu Ludności 2023 wg województw i płci.xlsx"
OUT_DIR = "../data"
YEAR = "2023"

SHEETS = {"Ogółem": "t", "Kobiety": "k", "Mężczyźni": "m"}

# Column index (0-based) -> voivodeship teryt, matching data/wojewodztwa.json's
# own 16 JPT_KOD_JE values exactly. Column 1 ("Polska ogółem") and columns
# 9-10 (mazowieckie's "w tym:" NUTS-2 sub-split) are deliberately excluded --
# see module docstring.
VOIVODESHIP_COLS = {
    2: "02",   # dolnośląskie
    3: "04",   # kujawsko-pomorskie
    4: "06",   # lubelskie
    5: "08",   # lubuskie
    6: "10",   # łódzkie
    7: "12",   # małopolskie
    8: "14",   # mazowieckie (whole voivodeship column, not the NUTS-2 split)
    11: "16",  # opolskie
    12: "18",  # podkarpackie
    13: "20",  # podlaskie
    14: "22",  # pomorskie
    15: "24",  # śląskie
    16: "26",  # świętokrzyskie
    17: "28",  # warmińsko-mazurskie
    18: "30",  # wielkopolskie
    19: "32",  # zachodniopomorskie
}

# Manual overrides for the two ambiguous/instructed cases -- see docstring.
# Row numbers are relative to each table's own start (table 1 starts at the
# sheet's row 7, table 2 at row 65 in "Ogółem"; identical structure repeats
# at fixed offsets in every sheet, discovered live per-sheet below rather
# than hardcoded, since blank rows can shift things slightly between sheets).


def hmm_to_min(v):
    if v is None:
        return None
    s = f"{v:.2f}"
    h_str, m_str = s.split(".")
    sign = -1 if v < 0 else 1
    return sign * (abs(int(h_str)) * 60 + int(m_str))


def is_all_caps_heading(pl):
    words = pl.split()
    return all(w.upper() == w or len(w) <= 1 for w in words) and any(c.isalpha() for c in pl)


def find_table_bounds(ws, max_row):
    """Locates the three stacked tables' header rows by their own title text."""
    bounds = {}
    for i in range(1, max_row + 1):
        b = ws[i][1].value
        if not isinstance(b, str):
            continue
        if b.startswith("PRZECIĘTNY CZAS TRWANIA"):
            bounds["trwania"] = i
        elif b.startswith("PRZECIĘTNY CZAS WYKONYWANIA"):
            bounds["wykonywania"] = i
        elif b.startswith("PRZECIĘTNY ODSETEK"):
            bounds["odsetek"] = i
    return bounds


def parse_table(ws, start, end, ncols, dojazdy_override=True):
    """Returns [(level, pl, teryt->minutes dict)] for L1/L2 rows only, plus
    the DOJAZDY override (skip the "w tym dojazdy i dojścia" line, promote
    its three children to L2) and the PRACA ZAWODOWA override (implicit --
    its only sub-line is "w tym praca główna...", already L3 by the normal
    "starts with w tym" rule, so no special-case needed for it)."""
    rows = []
    active_header = None
    last_level = None
    last_val0 = None  # column-0 (Polska ogółem) value, used only for the L1/L2 arithmetic checks
    skip_next_bare_header_reattach = False
    for i in range(start, end + 1):
        a = ws[i][0].value
        if a is None:
            continue
        parts = a.split("\n")
        pl = parts[0].strip()

        if pl == "w tym dojazdy i dojścia" and dojazdy_override:
            # Skip this L3 line entirely -- its children get promoted to L2
            # directly under DOJAZDY (last L1 seen), per explicit instruction.
            active_header = ("L1",)
            continue
        if pl == "w tym praca główna i dodatkowa":
            continue  # excluded entirely, per explicit instruction -- not promoted

        vals = {teryt: hmm_to_min(ws[i][col].value) for col, teryt in VOIVODESHIP_COLS.items()}
        v0 = hmm_to_min(ws[i][1].value)  # "Polska ogółem" column, for the arithmetic heuristic only

        if pl == "w tym:":
            if last_level == "L1":
                active_header = ("L1",)
            elif last_level == "L2":
                active_header = ("L2", last_val0, 0)
            else:
                active_header = ("deep",)
            continue

        is_l1 = is_all_caps_heading(pl)
        starts_w_tym = pl.lower().startswith("w tym")

        if is_l1:
            level = "L1"
            active_header = None
        elif starts_w_tym:
            level = "L3"
        elif active_header is not None and active_header[0] == "L1":
            level = "L2"
        elif active_header is not None and active_header[0] == "L2":
            _, parent_v0, running = active_header
            prospective = running + (v0 or 0)
            if v0 is not None and parent_v0 is not None and prospective <= parent_v0 + 2:
                level = "L3"
                reached = prospective >= parent_v0 - 1
                active_header = None if reached else ("L2", parent_v0, prospective)
            else:
                level = "L2"
                active_header = None
        elif active_header is not None and active_header[0] == "deep":
            level = "L3"
        else:
            level = "L2"

        last_level, last_val0 = level, v0
        if level in ("L1", "L2"):
            rows.append((level, pl, vals))
    return rows


def slugify(s):
    repl = {
        "ą": "a", "ć": "c", "ę": "e", "ł": "l", "ń": "n", "ó": "o", "ś": "s", "ź": "z", "ż": "z",
        "Ą": "A", "Ć": "C", "Ę": "E", "Ł": "L", "Ń": "N", "Ó": "O", "Ś": "S", "Ź": "Z", "Ż": "Z",
    }
    s = "".join(repl.get(c, c) for c in s)
    s = s.lower()
    out = []
    for c in s:
        out.append(c if c.isalnum() else "_")
    slug = "".join(out)
    while "__" in slug:
        slug = slug.replace("__", "_")
    return slug.strip("_")[:40]


def build_agegroups(rows):
    """[(key, pl_label, is_l1)] in source order, keys unique by construction
    (L2 keys are prefixed with their parent's own key)."""
    out = []
    parent_key = None
    for level, pl, _ in rows:
        if level == "L1":
            key = slugify(pl)
            parent_key = key
            out.append((key, pl, True))
        else:
            key = f"{parent_key}__{slugify(pl)}"
            out.append((key, pl, False))
    return out


def flat_text_lookup(ws, start, end):
    """{pl_text: teryt->minutes}, EVERY row in range (L1/L2/L3 alike, no
    classification at all) -- used only to pull table 2's VALUES by
    matching table 1's already-classified row text, since table 2's own
    values don't reliably partition/sum the way table 1's do (participation-
    time sub-items routinely exceed their own parent), which makes the
    arithmetic-tolerance heuristic parse_table() relies on invalid there.
    The two activity tables share an identical row-text sequence (same
    activities, just measured two different ways), so text matching is
    exact and doesn't need any arithmetic at all."""
    out = {}
    for i in range(start, end + 1):
        a = ws[i][0].value
        if a is None:
            continue
        pl = a.split("\n")[0].strip()
        if pl == "w tym:":
            continue
        out[pl] = {teryt: hmm_to_min(ws[i][col].value) for col, teryt in VOIVODESHIP_COLS.items()}
    return out


if __name__ == "__main__":
    wb = openpyxl.load_workbook(SRC, data_only=True)

    # Structure (row numbers, table order) confirmed identical across all
    # three sheets by locating each table's own header text live, not
    # assuming fixed row numbers.
    agegroups_master = None
    out = {}

    for sheet_name, sex_key in SHEETS.items():
        ws = wb[sheet_name]
        max_row = ws.max_row
        bounds = find_table_bounds(ws, max_row)
        t_start, w_start, o_start = bounds["trwania"], bounds["wykonywania"], bounds["odsetek"]

        # L1/L2 structure derived from table 1 ONLY -- trusted, since its
        # arithmetic-plausibility heuristic actually holds there (components
        # genuinely partition the parent). Table 2's structure is IDENTICAL
        # (same activities), so it's never independently re-classified --
        # see flat_text_lookup()'s own comment for why that would be wrong.
        rows_trwania = parse_table(ws, t_start + 1, w_start - 1, len(VOIVODESHIP_COLS))
        wykonywania_by_text = flat_text_lookup(ws, w_start + 1, o_start - 1)

        agegroups = build_agegroups(rows_trwania)
        if agegroups_master is None:
            agegroups_master = agegroups
        elif [a[0] for a in agegroups] != [a[0] for a in agegroups_master]:
            raise SystemExit(f"Category structure mismatch in sheet {sheet_name!r} -- expected identical L1/L2 keys across all three sheets")

        for (key, pl, _), (_, _, vals_trwania) in zip(agegroups, rows_trwania):
            vals_wykonywania = wykonywania_by_text.get(pl)
            if vals_wykonywania is None:
                raise SystemExit(f"Sheet {sheet_name!r}: no matching 'czas wykonywania' row found for {pl!r} -- table structures diverged unexpectedly")
            for measure_key, vals in (("trwania", vals_trwania), ("wykonywania", vals_wykonywania)):
                for teryt, minutes in vals.items():
                    slice_key = f"{key}__{measure_key}"
                    slot = out.setdefault(teryt, {}).setdefault(YEAR, {}).setdefault(slice_key, {"t": None, "m": None, "k": None})
                    slot[sex_key] = minutes

    path = os.path.join(OUT_DIR, "dobowy_budzet_czasu.json")
    with open(path, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, separators=(",", ":"))
    print(f"dobowy_budzet_czasu: {len(out)} regions -> {path}")

    # Dump the final PL ageGroups list (source order) for building the
    # variables.js entry -- not written to data/, just a build-time artifact.
    with open("/tmp/dobowy_budzet_czasu_agegroups.json", "w", encoding="utf-8") as f:
        json.dump([{"key": k, "label": pl, "isL1": is_l1} for k, pl, is_l1 in agegroups_master], f, ensure_ascii=False, indent=1)
    print(f"agegroups: {len(agegroups_master)} entries -> /tmp/dobowy_budzet_czasu_agegroups.json")
